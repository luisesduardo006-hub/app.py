timedeltaqlite3
import os
import random
import io
import pandas as pd
import mercadopago
from datetime import datetime, timedeltaapfrom flask import Flask, request, redirect, session, send_fileapp
appapp = Flask(__name__)
app.secret_key = 'SISTEMA_V10_PRO_FINAL_2026'

# --- CONFIGURACIÓN MERCADO PAGO ---
# Prioriza la variable de entorno de Render; si no existe, usa tu token directamente.
MP_TOKEN = os.environ.get("MP_ACCESS_TOKEN", "APP_USR-5698071543918489-041916-eb07a14c4a0b922a085b5e338cc595fe-3346852284")
sdk = mercadopago.SDK(MP_TOKEN)

# --- MOTOR DE DATOS ---
def get_db_connection():
    db_path = os.path.join(os.path.dirname(__file__), 'sistema_v9.db')
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn

def query_db(query, params=(), fetch=False):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(query, params)
        if fetch:
            res = cursor.fetchall()
            conn.close()
            return res
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Error DB: {e}")
        return []

def init_db():
    query_db('CREATE TABLE IF NOT EXISTS usuarios (clave TEXT PRIMARY KEY, nombre TEXT, rango TEXT, jefe TEXT)')
    query_db('CREATE TABLE IF NOT EXISTS productos (id INTEGER PRIMARY KEY AUTOINCREMENT, nombre TEXT, precio REAL, stock REAL, unidad TEXT, dueño TEXT)')
    query_db('CREATE TABLE IF NOT EXISTS ventas (id INTEGER PRIMARY KEY AUTOINCREMENT, total REAL, fecha TEXT, detalle TEXT, vendedor TEXT, dueño TEXT)')
    query_db('CREATE TABLE IF NOT EXISTS config (dueño TEXT PRIMARY KEY, empresa TEXT, whatsapp TEXT, estado TEXT, vencimiento TEXT)')
    query_db('CREATE TABLE IF NOT EXISTS pagos (id INTEGER PRIMARY KEY AUTOINCREMENT, concepto TEXT, monto REAL, fecha TEXT, responsable TEXT, dueño TEXT)')
    
    # Usuario Raíz (Admin Maestro)
    query_db("INSERT OR IGNORE INTO usuarios VALUES (?,?,?,?)", ('2026', 'Admin Maestro', 'Administrador', 'SISTEMA'))
    query_db("INSERT OR IGNORE INTO config VALUES (?,?,?,?,?)", ('2026', 'CENTRAL POS', '52', 'ACTIVO', '2030-01-01'))

# --- SEGURIDAD Y BLOQUEO AUTOMÁTICO ---
@app.before_request
def verificar_estatus():
    rutas_libres = ['/', '/auth', '/webhook_mp', '/generar_pago', '/health']
    if request.path in rutas_libres or 'static' in request.path:
        return
    
    if 'clv' in session and session['rango'] != 'Administrador':
        conf = query_db("SELECT estado, vencimiento FROM config WHERE dueño=?", (session['dueño'],), True)
        if conf:
            vencimiento = datetime.strptime(conf[0]['vencimiento'], '%Y-%m-%d')
            if datetime.now() > vencimiento or conf[0]['estado'] == 'SUSPENDIDO':
                return f'''{CSS}<div class="card" style="border-color:#f43f5e">
                    <h2 style="color:#f43f5e">SERVICIO SUSPENDIDO</h2>
                    <p style="text-align:center">La suscripción de <b>{session['dueño']}</b> ha vencido.</p>
                    <a href="/generar_pago" class="btn-nav" style="background:#009ee3; color:white; border:none; font-weight:800">💳 RENOVAR CON MERCADO PAGO</a>
                    <a href="/" class="btn-nav">Cerrar Sesión</a>
                </div>'''

# --- MERCADO PAGO: WEBHOOK Y PAGOS ---
@app.route('/health')
def health():
    return "OK", 200

@app.route('/webhook_mp', methods=['POST'])
def webhook_mp():
    if request.args.get('type') == 'payment':
        payment_id = request.args.get('data.id')
        payment_info = sdk.payment().get(payment_id)
        if payment_info["response"]["status"] == "approved":
            dueño_id = payment_info["response"]["external_reference"]
            nueva_fecha = (datetime.now() + timedelta(days=31)).strftime('%Y-%m-%d')
            query_db("UPDATE config SET estado='ACTIVO', vencimiento=? WHERE dueño=?", (nueva_fecha, dueño_id))
    return jsonify({"status": "ok"}), 200

@app.route('/generar_pago')
def generar_pago():
    if 'clv' not in session: return redirect('/')
    preference_data = {
        "items": [{"title": "Mensualidad Sistema POS", "quantity": 1, "unit_price": 450.00, "currency_id": "MXN"}],
        "external_reference": session['dueño'],
        "notification_url": f"https://{request.host}/webhook_mp",
        "back_urls": {"success": f"https://{request.host}/hub"}
    }
    res = sdk.preference().create(preference_data)
    return redirect(res["response"]["init_point"])

# --- ESTILOS CSS (GLASSMORPHISM) ---
CSS = '''
<style>
    @import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;600;800&display=swap');
    :root { --accent: #00f2fe; --bg: #0b0f1a; --glass: rgba(30, 41, 59, 0.7); --border: rgba(255, 255, 255, 0.1); }
    body { background: #0b0f1a; color: white; font-family: 'Plus Jakarta Sans', sans-serif; display: flex; flex-direction: column; align-items: center; padding: 20px; min-height: 100vh; }
    .card { background: var(--glass); backdrop-filter: blur(12px); border-radius: 20px; padding: 25px; width: 100%; max-width: 400px; border: 1px solid var(--border); margin-bottom: 15px; box-shadow: 0 10px 30px rgba(0,0,0,0.5); }
    h2 { background: linear-gradient(to right, #00f2fe, #4facfe); -webkit-background-clip: text; -webkit-text-fill-color: transparent; text-align: center; font-weight: 800; }
    input, select { background: rgba(0,0,0,0.3); border: 1px solid var(--border); color: white; padding: 12px; width: 100%; border-radius: 10px; margin-bottom: 10px; box-sizing: border-box; }
    button { background: linear-gradient(to right, #00f2fe, #4facfe); color: #0b0f1a; border: none; padding: 15px; width: 100%; border-radius: 10px; font-weight: 800; cursor: pointer; transition: 0.3s; }
    button:hover { opacity: 0.8; transform: scale(0.98); }
    .btn-nav { text-decoration: none; color: #94a3b8; display: block; text-align: center; padding: 12px; border-radius: 10px; border: 1px solid var(--border); margin-top: 10px; font-size: 14px; font-weight: 600; }
    table { width: 100%; font-size: 13px; border-collapse: collapse; }
    td { padding: 10px 0; border-bottom: 1px solid rgba(255,255,255,0.05); }
    .badge { background: var(--accent); color: black; padding: 3px 8px; border-radius: 5px; font-size: 10px; font-weight: 800; text-transform: uppercase; }
</style>
'''

# --- RUTAS DE NAVEGACIÓN ---
@app.route('/')
def login():
    session.clear()
    return f'{CSS}<div class="card"><h2>Acceso V10</h2><p style="text-align:center; color:#64748b; font-size:14px">Smart Business System</p><form action="/auth" method="POST"><input name="c" type="password" placeholder="Clave Operativa" required autofocus><button>INGRESAR</button></form></div>'

@app.route('/auth', methods=['POST'])
def auth():
    res = query_db("SELECT * FROM usuarios WHERE clave=?", (request.form['c'],), True)
    if res:
        session['clv'], session['user'], session['rango'] = res[0]['clave'], res[0]['nombre'], res[0]['rango']
        session['dueño'] = res[0]['jefe'] if res[0]['rango'] == 'Trabajador' else res[0]['clave']
        return redirect('/hub')
    return redirect('/')

@app.route('/hub')
def hub():
    if 'clv' not in session: return redirect('/')
    conf = query_db("SELECT empresa FROM config WHERE dueño=?", (session['dueño'],), True)
    empresa = conf[0]['empresa'] if conf else "SISTEMA"
    
    html = f'{CSS}<div class="card"><div style="text-align:center; margin-bottom:10px"><span class="badge">{session["rango"]}</span></div><h2>{empresa}</h2>'
    
    if session['rango'] in ['Trabajador', 'Dueño']:
        html += '<a href="/pos" class="btn-nav" style="background:var(--accent); color:black; border:none; font-weight:800">🛒 VENTAS / CAJA</a>'
        html += '<a href="/proveedores" class="btn-nav">🚚 PAGO A PROVEEDORES</a>'
        html += '<a href="/inventario" class="btn-nav">📦 INVENTARIO</a>'
    
    if session['rango'] == 'Dueño':
        html += '<div style="margin-top:20px; border-top:1px solid var(--border); padding-top:10px"></div>'
        html += '<a href="/corte" class="btn-nav" style="border-color:var(--accent); color:var(--accent)">📊 CORTE DE CAJA</a>'
        html += '<a href="/gestion_personal" class="btn-nav">👥 GESTIÓN DE PERSONAL</a>'
        html += '<a href="/ajustes" class="btn-nav">⚙️ CONFIGURACIÓN</a>'
        
    if session['rango'] == 'Administrador':
        html += '<a href="/gestion_dueños" class="btn-nav" style="background:#f59e0b; color:black; border:none; font-weight:800">🏢 PANEL DE ADMINISTRADOR</a>'

    html += '<a href="/" class="btn-nav" style="color:#f43f5e; border-color:rgba(244,63,94,0.2); margin-top:30px">Cerrar Sesión Segura</a></div>'
    return html

# --- VENTAS Y BUSCADOR ---
@app.route('/pos')
def pos():
    prods = query_db("SELECT * FROM productos WHERE dueño=?", (session['dueño'],), True)
    carro = session.get('carro', [])
    total = sum(i['s'] for i in carro)
    
    l_bus = "".join([f'<div style="padding:12px; border-bottom:1px solid var(--border); cursor:pointer" onclick="document.getElementById(\'pid\').value=\'{p["id"]}\'; document.getElementById(\'q\').value=\'{p["nombre"]}\'; this.parentElement.style.display=\'none\'">{p["nombre"]} - <span style="color:var(--accent)">${p["precio"]}</span> <small>({p["stock"] if p["stock"] is not None else "∞"})</small></div>' for p in prods])
    
    return f'''{CSS}<div class="card"><h2>Terminal de Venta</h2>
        <table>{"".join([f"<tr><td>{i['n']}</td><td style='text-align:right'>${i['s']}</td></tr>" for i in carro])}</table>
        <div style="background:rgba(0,242,254,0.1); padding:15px; border-radius:15px; text-align:center; margin:15px 0">
            <small>TOTAL A PAGAR</small><h2 style="margin:0; font-size:32px">${total}</h2>
        </div>
        <form action="/pagar" method="POST"><input name="tel" placeholder="WhatsApp Cliente (ej: 521...)" required><button>FINALIZAR VENTA</button></form>
        <div style="margin-top:20px; border-top:1px dashed var(--border); padding-top:15px">
            <input id="q" onkeyup="this.nextElementSibling.style.display='block'" placeholder="🔍 Buscar producto...">
            <div style="display:none; background:#1e293b; border-radius:10px; position:absolute; width:85%; z-index:10; border:1px solid var(--accent); max-height:200px; overflow-y:auto">{l_bus}</div>
            <form action="/add_carro" method="POST" style="display:flex; gap:10px; margin-top:10px">
                <input type="hidden" name="id" id="pid">
                <input name="val" type="number" step="0.01" placeholder="Cant/Monto" required>
                <button style="width:60px">+</button>
            </form>
        </div><a href="/hub" class="btn-nav">Regresar al Menú</a></div>'''

@app.route('/add_carro', methods=['POST'])
def add_carro():
    p = query_db("SELECT * FROM productos WHERE id=?", (request.form['id'],), True)[0]
    val = float(request.form['val'])
    sub = val if p['unidad'] == 'Kg' else round(val * p['precio'], 2)
    carro = session.get('carro', []); carro.append({'id': p['id'], 'n': p['nombre'], 'c': val/p['precio'] if p['unidad']=='Kg' else val, 's': sub})
    session['carro'] = carro
    return redirect('/pos')

@app.route('/pagar', methods=['POST'])
def pagar():
    carro = session.get('carro', [])
    if not carro: return redirect('/pos')
    
    conf = query_db("SELECT empresa FROM config WHERE dueño=?", (session['dueño'],), True)
    empresa = conf[0]['empresa'] if conf else "MI NEGOCIO"
    total = sum(i['s'] for i in carro)
    
    ticket_prod = ""
    for i in carro:
        query_db("UPDATE productos SET stock = stock - ? WHERE id = ? AND stock IS NOT NULL", (i['c'], i['id']))
        query_db("INSERT INTO ventas (total, fecha, detalle, vendedor, dueño) VALUES (?,?,?,?,?)", (i['s'], datetime.now().strftime("%H:%M"), i['n'], session['user'], session['dueño']))
        ticket_prod += f"• {i['n']} ({i['c']}) ${i['s']}%0A"

    mensaje = f"🧾 *TICKET DE VENTA*%0A"
    mensaje += f"------------------------------%0A"
    mensaje += f"🏪 *{empresa.upper()}*%0A"
    mensaje += f"👤 Atendió: {session['user']}%0A"
    mensaje += f"------------------------------%0A"
    mensaje += ticket_prod
    mensaje += f"------------------------------%0A"
    mensaje += f"💰 *TOTAL: ${total}*%0A%0A"
    mensaje += f"🙏 ¡Gracias por su compra!"

    session['carro'] = []
    return redirect(f"https://api.whatsapp.com/send?phone={request.form['tel']}&text={mensaje}")
    

# --- CORTE E INVENTARIO ---
@app.route('/corte')
def corte():
    v_l = query_db("SELECT * FROM ventas WHERE dueño=? ORDER BY id DESC", (session['dueño'],), True)
    p_l = query_db("SELECT * FROM pagos WHERE dueño=? ORDER BY id DESC", (session['dueño'],), True)
    t_v, t_p = sum(v['total'] for v in v_l), sum(p['monto'] for p in p_l)
    return f'''{CSS}<div class="card"><h2>Corte Diario</h2>
        <div style="text-align:center; background:rgba(34,197,94,0.1); padding:20px; border-radius:20px; border:1px solid #22c55e; margin-bottom:15px">
            <small>NETO EN CAJA</small><h1 style="margin:5px 0">${t_v - t_p}</h1>
            <p style="font-size:12px; color:#94a3b8">Ventas: ${t_v} | Gastos: -${t_p}</p>
        </div>
        <a href="/exportar_excel" class="btn-nav" style="background:#1d6f42; color:white; border:none">📑 DESCARGAR REPORTE EXCEL</a>
        <h3 style="font-size:14px; margin-top:20px; color:var(--accent)">Últimos Movimientos</h3>
        <table>{"".join([f"<tr><td>{v['fecha']}</td><td>{v['detalle']}</td><td style='color:#22c55e'>+${v['total']}</td></tr>" for v in v_l[:10]])}</table>
        <a href="/hub" class="btn-nav">Volver</a></div>'''

@app.route('/exportar_excel')
def exportar_excel():
    conn = get_db_connection()
    dueño_id = session['dueño']
    
    # 1. Obtener Ventas (Incluyendo ID como No. Ticket)
    query_v = """
    SELECT id as 'TICKET', date('now') as FECHA, fecha as HORA, vendedor as 'RESPONSABLE', 
    'VENTA' as TIPO, detalle as 'CONCEPTO / PRODUCTO', 
    total as 'ENTRADA (+)', 0.0 as 'SALIDA (-)', 'CLIENTE' as 'DESTINATARIO'
    FROM ventas WHERE dueño=?
    """
    df_v = pd.read_sql_query(query_v, conn, params=(dueño_id,))
    
    # 2. Obtener Pagos
    query_p = """
    SELECT id as 'TICKET', date('now') as FECHA, fecha as HORA, responsable as 'RESPONSABLE', 
    'PAGO PROV' as TIPO, concepto as 'CONCEPTO / PRODUCTO', 
    0.0 as 'ENTRADA (+)', monto as 'SALIDA (-)', 'PROVEEDOR' as 'DESTINATARIO'
    FROM pagos WHERE dueño=?
    """
    df_p = pd.read_sql_query(query_p, conn, params=(dueño_id,))
    
    # Si no hay nada, avisamos para no borrar tablas vacías
    if df_v.empty and df_p.empty:
        return f'{CSS}<div class="card"><h2>Sin datos</h2><p>No hay movimientos hoy.</p><a href="/hub" class="btn-nav">Volver</a></div>'

    # Unir y ordenar
    df_final = pd.concat([df_v, df_p]).sort_values(by='HORA')

    # --- BORRADO TRAS GENERAR EL REPORTE ---
    query_db("DELETE FROM ventas WHERE dueño=?", (dueño_id,))
    query_db("DELETE FROM pagos WHERE dueño=?", (dueño_id,))

    conn.close()

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_final.to_excel(writer, index=False, sheet_name='Cierre de Caja')
        ws = writer.sheets['Cierre de Caja']
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # Estilos Pasteles
        color_header = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
        color_venta = PatternFill(start_color="E2F9E1", end_color="E2F9E1", fill_type="solid")
        color_pago = PatternFill(start_color="FFF1F0", end_color="FFF1F0", fill_type="solid")
        font_white = Font(bold=True, color="FFFFFF")

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                if cell.row == 1:
                    cell.fill = color_header
                    cell.font = font_white
                    cell.alignment = Alignment(horizontal="center")
                else:
                    # Columna E es 'TIPO'
                    tipo = ws[f"E{cell.row}"].value
                    cell.fill = color_venta if tipo == 'VENTA' else color_pago

        for col in ws.columns:
            max_len = max([len(str(cell.value)) for cell in col])
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                     as_attachment=True, download_name=f"CIERRE_{datetime.now().strftime('%d_%m_%H%M')}.xlsx")
    

@app.route('/inventario')
def inventario():
    ps = query_db("SELECT * FROM productos WHERE dueño=?", (session['dueño'],), True)
    html = f'{CSS}<div class="card"><h2>Inventario</h2><table style="margin-bottom:20px"><tr><th>Producto</th><th>Precio</th><th>Stock</th></tr>{"".join([f"<tr><td>{p['nombre']}</td><td>${p['precio']}</td><td>{p['stock'] if p['stock'] is not None else '∞'}</td></tr>" for p in ps])}</table>'
    if session['rango'] == 'Dueño':
        html += '<div style="border-top:1px solid var(--border); padding-top:15px"><h3>Añadir Nuevo</h3><form action="/save_prod" method="POST"><input name="n" placeholder="Nombre del producto"><input name="p" type="number" step="0.01" placeholder="Precio"><input name="s" type="number" placeholder="Stock (Vacío para ∞)"><select name="u"><option value="Pza">Por Pieza</option><option value="Kg">Por Kilo (Monto libre)</option></select><button>REGISTRAR PRODUCTO</button></form></div>'
    html += '<a href="/hub" class="btn-nav">Volver</a></div>'
    return html

@app.route('/save_prod', methods=['POST'])
def save_prod():
    query_db("INSERT INTO productos (nombre, precio, stock, unidad, dueño) VALUES (?,?,?,?,?)", (request.form['n'], request.form['p'], request.form['s'] or None, request.form['u'], session['dueño']))
    return redirect('/inventario')

# --- GASTOS Y PERSONAL ---
@app.route('/proveedores')
def proveedores():
    pagos = query_db("SELECT * FROM pagos WHERE dueño=? ORDER BY id DESC", (session['dueño'],), True)
    return f'{CSS}<div class="card"><h2>Gastos / Proveedores</h2><form action="/add_pago" method="POST"><input name="con" placeholder="Concepto (ej: Sabritas)"><input name="mon" type="number" step="0.01" placeholder="Monto Pagado"><button style="background:#f59e0b; color:black">REGISTRAR SALIDA</button></form><h3>Gastos de Hoy</h3><table>{"".join([f"<tr><td>{p['concepto']}</td><td style='color:#f43f5e; font-weight:bold'>-${p['monto']}</td></tr>" for p in pagos])}</table><a href="/hub" class="btn-nav">Volver</a></div>'

@app.route('/add_pago', methods=['POST'])
def add_pago():
    query_db("INSERT INTO pagos (concepto, monto, fecha, responsable, dueño) VALUES (?,?,?,?,?)", (request.form['con'], request.form['mon'], datetime.now().strftime("%H:%M"), session['user'], session['dueño']))
    return redirect('/proveedores')

@app.route('/gestion_personal')
def gestion_personal():
    lista = query_db("SELECT * FROM usuarios WHERE jefe=?", (session['clv'],), True)
    return f'{CSS}<div class="card"><h2>Mi Personal</h2><table>{"".join([f"<tr><td>{u['nombre']}</td><td>Clave: <code>{u['clave']}</code></td></tr>" for u in lista])}</table><hr><form action="/crear_usuario" method="POST"><input name="n" placeholder="Nombre del Empleado"><button>CREAR ACCESO TRABAJADOR</button></form><a href="/hub" class="btn-nav">Volver</a></div>'

@app.route('/crear_usuario', methods=['POST'])
def crear_usuario():
    query_db("INSERT INTO usuarios VALUES (?,?,?,?)", (f"TR-{random.randint(1000,9999)}", request.form['n'], 'Trabajador', session['clv']))
    return redirect('/gestion_personal')

@app.route('/ajustes')
def ajustes():
    c = query_db("SELECT * FROM config WHERE dueño=?", (session['dueño'],), True)[0]
    return f'{CSS}<div class="card"><h2>Configuración</h2><form action="/save_config" method="POST"><input name="e" value="{c["empresa"]}" placeholder="Nombre del Negocio"><input name="w" value="{c["whatsapp"]}" placeholder="WhatsApp Reportes"><button>GUARDAR CAMBIOS</button></form><a href="/hub" class="btn-nav">Volver</a></div>'

@app.route('/save_config', methods=['POST'])
def save_config():
    query_db("UPDATE config SET empresa=?, whatsapp=? WHERE dueño=?", (request.form['e'], request.form['w'], session['dueño']))
    return redirect('/hub')

# --- PANEL ADMIN (TÚ) ---
@app.route('/gestion_dueños')
def gestion_dueños():
    if session.get('rango') != 'Administrador': return redirect('/')
    dueños = query_db("SELECT u.nombre, u.clave, c.empresa, c.vencimiento, c.estado FROM usuarios u JOIN config c ON u.clave=c.dueño WHERE u.rango='Dueño'", (), True)
    html = f'{CSS}<div class="card"><h2>Control de Dueños</h2><table>'
    for d in dueños:
        html += f"<tr><td><b>{d['empresa']}</b><br><small>Vence: {d['vencimiento']}</small></td><td style='text-align:right'><span class='badge'>{d['estado']}</span></td></tr>"
    html += '</table><h3>Registrar Nuevo Dueño</h3><form action="/crear_dueño" method="POST"><input name="n" placeholder="Nombre Dueño"><input name="e" placeholder="Empresa"><input name="c" placeholder="Clave Acceso"><button>DAR DE ALTA</button></form><a href="/hub" class="btn-nav">Volver</a></div>'
    return html

@app.route('/crear_dueño', methods=['POST'])
def crear_dueño():
    clv = request.form['c']
    vence = (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d')
    query_db("INSERT INTO usuarios VALUES (?,?,?,?)", (clv, request.form['n'], 'Dueño', clv))
    query_db("INSERT INTO config VALUES (?,?,?,?,?)", (clv, request.form['e'], '52', 'ACTIVO', vence))
    return redirect('/gestion_dueños')

if __name__ == "__main__":
    init_db()
    port = int(os.environ.get("PORT", 10000))
    app.run(host='0.0.0.0', port=port)
    
